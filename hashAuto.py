import os
import hashlib
import openpyxl
from tkinter import Tk, filedialog, messagebox

def select_directory():
    root = Tk()
    root.withdraw()
    directory_path = filedialog.askdirectory()
    root.destroy()
    return directory_path

def calculate_sha256(file_path):
    with open(file_path, 'rb') as file:
        file_data = file.read()
        hash_object = hashlib.sha256(file_data)
        return hash_object.hexdigest()

def process_directory(directory_path):
    hashes = []

    for root, _, files in os.walk(directory_path):
        for file in files:
            file_path = os.path.join(root, file)
            hash_value = calculate_sha256(file_path)
            relative_path = os.path.relpath(file_path, directory_path)
            hashes.append({'file': relative_path, 'hash': hash_value})

    return hashes

def export_to_excel(hashes, directory_path):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Hashes'

    worksheet['A1'] = 'Arquivo'
    worksheet['B1'] = 'SHA256'

    worksheet.column_dimensions['A'].width = 100
    worksheet.column_dimensions['B'].width = 100

    for index, hash_data in enumerate(hashes, start=2):
        worksheet.cell(row=index, column=1, value=hash_data['file'])
        worksheet.cell(row=index, column=2, value=hash_data['hash'])

    output_file = os.path.join(directory_path, 'Hashes.xlsx')

    workbook.save(output_file)
    messagebox.showinfo('Exportação Concluída', f'Arquivo Excel exportado com sucesso: {output_file}')

def browse_directory():
    directory_path = select_directory()
    if directory_path:
        hashes = process_directory(directory_path)
        export_to_excel(hashes, directory_path)

browse_directory()
